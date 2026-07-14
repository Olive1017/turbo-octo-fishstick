"""
归类模块：按车队将下载的 PDF 移动到对应车队子文件夹
"""
import os
import shutil
import glob
from datetime import datetime
from openpyxl import load_workbook

ILLEGAL_CHARS = '\\/:*?"<>|'


def _safe_name(name: str) -> str:
    """清洗车队名中的非法文件名字符"""
    if not name:
        return "未分类"
    for ch in ILLEGAL_CHARS:
        name = name.replace(ch, "_")
    return name.strip() or "未分类"


def find_excel(save_dir):
    """
    在保存目录下查找唯一的 xlsx。
    唯一时返回路径；0 个或多个返回 None（由 GUI 弹窗让用户选）。
    """
    files = [
        f for f in glob.glob(os.path.join(save_dir, "*.xlsx"))
        if not os.path.basename(f).startswith("~$")
    ]
    return files[0] if len(files) == 1 else None


def classify(save_dir, excel_path):
    """执行归类，返回统计 dict"""
    wb = load_workbook(excel_path)
    ws = wb.active

    header = {}
    for idx, cell in enumerate(ws[1]):
        if cell.value is not None:
            header[str(cell.value).strip()] = idx

    if "SAP订单号" not in header or "车队" not in header:
        raise Exception("Excel 表头未找到 'SAP订单号' 或 '车队' 列，请检查表头")

    sap_col = header["SAP订单号"]
    fleet_col = header["车队"]

    mapping = {}
    for row in ws.iter_rows(min_row=2):
        sap = str(row[sap_col].value).strip() if row[sap_col].value is not None else ""
        fleet = str(row[fleet_col].value).strip() if row[fleet_col].value is not None else ""
        if sap:
            mapping[sap] = fleet if fleet else "未分类"
    print(f"Excel 共读取 {len(mapping)} 条订单记录")

    pdf_files = {}
    for f in os.listdir(save_dir):
        full = os.path.join(save_dir, f)
        if os.path.isfile(full) and f.lower().endswith(".pdf"):
            pdf_files[os.path.splitext(f)[0]] = f
    print(f"根目录共发现 {len(pdf_files)} 个 PDF 文件")

    moved = 0
    uncategorized = 0
    uncat_list = []

    for order, filename in pdf_files.items():
        fleet = mapping.get(order, "未分类")
        if fleet == "未分类":
            uncategorized += 1
            uncat_list.append(order)

        target_dir = os.path.join(save_dir, _safe_name(fleet))
        os.makedirs(target_dir, exist_ok=True)

        src = os.path.join(save_dir, filename)
        dst = os.path.join(target_dir, filename)
        if os.path.exists(dst):
            os.remove(dst)
        shutil.move(src, dst)
        moved += 1

    missing_list = [o for o in mapping if o not in pdf_files]

    status_col = ws.max_column + 1
    ws.cell(row=1, column=status_col, value="归类状态")
    for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
        sap = str(row[sap_col].value).strip() if row[sap_col].value is not None else ""
        if not sap:
            continue
        ws.cell(row=i, column=status_col, value="已归类" if sap in pdf_files else "缺PDF")

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    result_path = os.path.join(save_dir, f"归类结果_{timestamp}.xlsx")
    wb.save(result_path)
    print(f"归类结果已另存为: {result_path}")

    if uncat_list:
        print(f"\n【未分类】以下 {len(uncat_list)} 个 PDF 在 Excel 中查不到车队，已移入“未分类”：")
        for o in uncat_list:
            print(f"  - {o}")
    if missing_list:
        print(f"\n【缺PDF】以下 {len(missing_list)} 个订单在 Excel 中存在但缺对应 PDF：")
        for o in missing_list:
            print(f"  - {o}")

    return {"moved": moved, "uncategorized": uncategorized, "missing": len(missing_list)}
